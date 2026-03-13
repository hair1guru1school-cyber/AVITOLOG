import openpyxl
wb = openpyxl.load_workbook(r'c:\Users\shink\Downloads\CRMM.xlsx', data_only=True)
out = []
out.append('=== SHEET NAMES ===')
out.append(str(wb.sheetnames))
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    out.append('\n=== SHEET: ' + sheet_name + ' ===')
    out.append('Dimensions: ' + str(ws.dimensions))
    out.append('Max row: ' + str(ws.max_row) + ', Max column: ' + str(ws.max_column))
    out.append('\n--- First 20 rows (all columns) ---')
    for r in range(1, min(21, ws.max_row+1)):
        row_vals = []
        for c in range(1, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            row_vals.append(str(val)[:50] if val is not None else '')
        out.append('Row ' + str(r) + ': ' + str(row_vals))
    out.append('\n--- Column widths ---')
    for c in range(1, ws.max_column+1):
        width = ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width
        out.append('  Col ' + str(c) + ' (' + openpyxl.utils.get_column_letter(c) + '): width=' + str(width))
    out.append('\n--- Row heights (first 30 rows) ---')
    for r in range(1, min(31, ws.max_row+1)):
        h = ws.row_dimensions[r].height
        out.append('  Row ' + str(r) + ': height=' + str(h))
with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(out))
print('Done')
